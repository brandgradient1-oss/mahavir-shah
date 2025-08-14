import os
import re
import uuid
import json
import io
import time
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any, Tuple

import pandas as pd
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from fastapi import FastAPI, APIRouter, UploadFile, File, Form, HTTPException, Query
from fastapi.responses import JSONResponse, FileResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from starlette.middleware.cors import CORSMiddleware

# ---------- Env & Config ----------
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ.get('DB_NAME', 'test_database')
client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY')
GOOGLE_API_KEY = os.environ.get('GOOGLE_API_KEY')
CSE_ID = os.environ.get('CSE_ID')
BING_SEARCH_KEY = os.environ.get('BING_SEARCH_KEY')

try:
    from google import genai
except Exception as e:
    genai = None
    logging.warning('google-genai not installed or failed to import. Install google-genai in requirements.txt')

app = FastAPI()
api_router = APIRouter(prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ---------- Models ----------
class StatusCheck(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=datetime.utcnow)

class StatusCheckCreate(BaseModel):
    client_name: str

class ScrapeMode(str):
    REALTIME = 'realtime'
    DEEP = 'deep'

class ScrapeRequest(BaseModel):
    url: str
    mode: str = Field(default=ScrapeMode.REALTIME)

class NameGeoRequest(BaseModel):
    company_name: str
    geography: Optional[str] = ""
    mode: str = Field(default=ScrapeMode.REALTIME)

class ScrapeResult(BaseModel):
    job_id: str
    status: str
    data: Dict[str, Any]
    excel_path: Optional[str]
    created_at: datetime

# ---------- Utilities ----------
SOCIAL_DOMAINS = [
    'linkedin.com', 'facebook.com', 'instagram.com', 'twitter.com', 'x.com', 'youtube.com', 't.me'
]

MOTH_HEADERS = [
    "Company Name","Website","Industry","Description","Services","Address","Country","State","City","Postal Code","Phone","Email","Social Media Links","Founders/Key People","Verification Status","Scraped At"
]

STRIP_SUFFIXES = [
    'private limited','pvt ltd','pvt. ltd.','limited','ltd','llp','inc','inc.','co','co.','company','solutions','solution','technologies','technology','tech','labs','studio','studios','group','services','service','global','international'
]

def clean_text(s: Optional[str]) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", " ", BeautifulSoup(s, 'lxml').get_text(" ").strip())


def find_emails(text: str) -> List[str]:
    if not text:
        return []
    pattern = r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}"
    return sorted(set(re.findall(pattern, text)))


def find_phones(text: str) -> List[str]:
    if not text:
        return []
    pattern = r"\+?\d[\d\s().-]{6,}\d"
    return sorted(set(re.findall(pattern, text)))


def absolute_url(base: str, link: str) -> str:
    try:
        return requests.compat.urljoin(base, link)
    except Exception:
        return link


def normalize_url(u: str) -> str:
    if not u:
        return u
    u = u.strip()
    if not u.startswith("http://") and not u.startswith("https://"):
        u = "https://" + u
    p = requests.utils.urlparse(u)
    if not p.netloc and p.path:
        u = f"https://{p.path}"
    return u


def try_fetch(url: str, headers: Dict[str, str]) -> Optional[requests.Response]:
    candidates = []
    base = normalize_url(url)
    p = requests.utils.urlparse(base)
    host = p.netloc
    no_www = host.replace("www.", "")
    with_www = host if host.startswith("www.") else f"www.{host}"
    candidates.append(base)
    candidates.append(base.replace("http://", "https://"))
    candidates.append(base.replace("https://", "http://"))
    candidates.append(f"https://{with_www}{p.path or ''}")
    candidates.append(f"http://{with_www}{p.path or ''}")
    candidates.append(f"https://{no_www}{p.path or ''}")
    candidates = list(dict.fromkeys(candidates))

    for c in candidates:
        try:
            r = requests.get(c, timeout=15, headers=headers, allow_redirects=True)
            if (200 <= r.status_code < 400) or (r.status_code in (401,403) and len(r.text) > 500):
                return r
        except Exception as e:
            logger.warning(f"Fetch attempt failed {c}: {e}")
    return None


def crawl_site(url: str, mode: str = ScrapeMode.REALTIME, max_pages: int = 10) -> Dict[str, Any]:
    visited = set()
    to_visit = [normalize_url(url)]
    pages = {}
    count_limit = 2 if mode == ScrapeMode.REALTIME else max_pages
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"}

    while to_visit and len(visited) < count_limit:
        current = to_visit.pop(0)
        if current in visited:
            continue
        try:
            r = try_fetch(current, headers)
            if not r:
                visited.add(current)
                continue
            html = r.text
            pages[current] = html
            visited.add(current)

            soup = BeautifulSoup(html, 'lxml')
            for a in soup.find_all('a', href=True):
                href = a['href']
                if any(dom in href for dom in SOCIAL_DOMAINS):
                    full = absolute_url(current, href)
                    pages[full] = ''
                else:
                    absu = absolute_url(current, href)
                    try:
                        if requests.utils.urlparse(absu).netloc.endswith(requests.utils.urlparse(current).netloc.replace('www.', '')):
                            if absu not in visited and len(to_visit) < 50:
                                to_visit.append(absu)
                    except Exception:
                        pass
        except Exception as e:
            logger.warning(f"Failed to fetch {current}: {e}")
            visited.add(current)
            continue
    return pages

# ---------- AI extraction ----------

def init_gemini_client():
    if genai is None:
        raise HTTPException(status_code=500, detail='google-genai not installed')
    if not GEMINI_API_KEY:
        raise HTTPException(status_code=400, detail='GEMINI_API_KEY missing in backend/.env')
    try:
        return genai.Client(api_key=GEMINI_API_KEY)
    except Exception as e:
        logger.error(f"Gemini client init failed: {e}")
        raise HTTPException(status_code=500, detail=f"Gemini init failed: {e}")


def _strip_suffixes(name: str) -> str:
    n = name.lower()
    for s in STRIP_SUFFIXES:
        n = re.sub(rf"\b{s}\b", " ", n)
    n = re.sub(r"[^a-z0-9 ]", " ", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def _compose_candidates(name: str, geography: str = "") -> List[str]:
    base = _strip_suffixes(name)
    parts = [p for p in base.split(' ') if p]
    if not parts:
        return []
    compact = ''.join(parts)
    dashed = '-'.join(parts)
    tlds = ['com','co','io','ai','net','org','in','co.in','info','biz','tech']
    hosts = set()
    for t in tlds:
        hosts.add(f"{compact}.{t}")
        if len(parts) > 1:
            hosts.add(f"{dashed}.{t}")
    return list(hosts)


def _fetch_meta_for_hosts(hosts: List[str]) -> List[Dict[str, Any]]:
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"}
    out = []
    for h in hosts[:25]:
        for scheme in ["https://", "http://"]:
            url = scheme + h
            try:
                r = try_fetch(url, headers)
                if not r:
                    continue
                html = r.text
                soup = BeautifulSoup(html, 'lxml')
                title = (soup.title.string if soup and soup.title else '') or ''
                md = soup.find('meta', attrs={'name':'description'}) or soup.find('meta', attrs={'property':'og:description'})
                desc = md['content'] if md and md.get('content') else ''
                out.append({"host": h, "url": url, "title": clean_text(title)[:200], "desc": clean_text(desc)[:400]})
                break
            except Exception:
                continue
    return out


def gemini_select_official(company: str, geography: str, candidates: List[Dict[str, Any]]) -> Optional[str]:
    if not candidates:
        return None
    client = init_gemini_client()
    summary = json.dumps(candidates)[:12000]
    instruction = (
        "You are an AI that selects the most likely official website for a company based on candidates.\n"
        "Return strict JSON: {\"domain\": string, \"confidence\": number, \"rationale\": string}.\n"
        "Consider exact name match, brand signals in title/desc, and ignore directories like careers/help."
    )
    user_text = f"Company: {company}\nLocation hint: {geography}\nCandidates(JSON):\n{summary}\nReturn only JSON."
    try:
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[{"role":"user","parts":[{"text": instruction + "\n\n" + user_text}]}]
        )
        raw = getattr(resp,'text','') or ''
        m = re.search(r"\{[\s\S]*\}", raw)
        if not m:
            return candidates[0]['url']
        js = json.loads(m.group(0))
        domain_url = js.get('domain') or js.get('url')
        if domain_url:
            # normalize to url
            if domain_url.startswith('http'):
                return domain_url
            return 'https://' + domain_url
        return candidates[0]['url']
    except Exception as e:
        logger.warning(f"Gemini select official failed: {e}")
        return candidates[0]['url']


def resolve_site_via_ai(company: str, geography: str = "") -> str:
    hosts = _compose_candidates(company, geography)
    candidates = _fetch_meta_for_hosts(hosts)
    if candidates:
        chosen = gemini_select_official(company, geography, candidates)
        if chosen:
            return chosen
    # last resort: use guessed .com
    if hosts:
        return "https://" + hosts[0]
    raise HTTPException(status_code=400, detail='Unable to resolve official website')


def gemini_extract(company_url: str, pages: Dict[str, str]) -> Dict[str, Any]:
    client = init_gemini_client()
    combined = []
    take = 2
    for i, (u, html) in enumerate(pages.items()):
        if i >= take:
            break
        text = clean_text(html)
        combined.append(f"URL: {u}\n{text[:8000]}")
    prompt_text = "\n\n".join(combined)[:16000]

    system_prompt = (
        "You are an AI data extractor. Input is raw text from a company website. Return JSON exactly in this schema with English fields: {\n"
        "\"Company Name\": \"\",\n"
        "\"Website\": \"\",\n"
        "\"Industry\": \"\",\n"
        "\"Description\": \"\",\n"
        "\"Services\": \"\",\n"
        "\"Address\": \"\",\n"
        "\"Country\": \"\",\n"
        "\"State\": \"\",\n"
        "\"City\": \"\",\n"
        "\"Postal Code\": \"\",\n"
        "\"Phone\": \"\",\n"
        "\"Email\": \"\",\n"
        "\"Social Media Links\": \"\",\n"
        "\"Founders/Key People\": \"\",\n"
        "\"Verification Status\": \"UNVERIFIED\"\n}"
    )

    try:
        user_text = f"Instructions:\n{system_prompt}\n\nWebsite content:\n{prompt_text}\n\nFollow the instructions strictly."
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[{"role":"user","parts":[{"text": user_text}]}]
        )
        raw = getattr(resp, 'text', '') or ''
        m = re.search(r"\{[\s\S]*\}", raw)
        if m:
            data = json.loads(m.group(0))
        else:
            raise ValueError("no-json")
    except Exception as e:
        logger.warning(f"Gemini extraction failed or returned non-JSON, falling back. Reason: {e}")
        first_html = next(iter(pages.values())) if pages else ''
        soup = BeautifulSoup(first_html, 'lxml') if first_html else None
        title = (soup.title.string if soup and soup.title else '') or ''
        meta_desc = ''
        if soup:
            md = soup.find('meta', attrs={'name':'description'}) or soup.find('meta', attrs={'property':'og:description'})
            if md and md.get('content'):
                meta_desc = md['content']
        para = ''
        if soup:
            p = soup.find('p')
            para = p.get_text(" ").strip() if p else ''
        data = {
            "Company Name": title.strip()[:200],
            "Website": company_url,
            "Industry": "",
            "Description": (meta_desc or para)[:1000],
            "Services": "",
            "Address": "",
            "Country": "",
            "State": "",
            "City": "",
            "Postal Code": "",
            "Phone": "",
            "Email": "",
            "Social Media Links": "",
            "Founders/Key People": "",
            "Verification Status": "UNVERIFIED"
        }

    if not data.get("Website"):
        data["Website"] = company_url

    html_concat = "\n".join(pages.values())
    emails = find_emails(html_concat)
    phones = find_phones(html_concat)
    if emails and not data.get("Email"):
        data["Email"] = ", ".join(emails[:3])
    if phones and not data.get("Phone"):
        data["Phone"] = ", ".join(phones[:3])

    socials = sorted({u for u in pages.keys() if any(dom in u for dom in SOCIAL_DOMAINS)})
    if socials:
        existing = data.get("Social Media Links", "")
        merged = ", ".join(sorted(set([*([s.strip() for s in existing.split(',') if s.strip()]), *socials])))
        data["Social Media Links"] = merged

    data["Verification Status"] = "UNVERIFIED"
    return data

# ---------- AI-only contact verification ----------

def _norm_phone(p: str) -> str:
    if not p:
        return ''
    p = re.sub(r"[^+\d]", "", p)
    if p.startswith('+'):
        core = re.sub(r"\D", "", p)
        return '+' + core
    return re.sub(r"\D", "", p)


def _domain_from_url(u: str) -> str:
    try:
        host = requests.utils.urlparse(normalize_url(u)).netloc.lower()
        return host.replace('www.', '')
    except Exception:
        return ''


def ai_verify_contacts(data: Dict[str, Any], pages: Dict[str, str], site_url: str) -> Tuple[Dict[str, Any], str]:
    site_domain = _domain_from_url(site_url)
    per_url_emails: Dict[str, set] = {}
    per_url_phones: Dict[str, set] = {}

    for u, html in pages.items():
        if not html:
            continue
        es = find_emails(html)
        ps = find_phones(html)
        if es:
            per_url_emails[u] = set(es)
        if ps:
            per_url_phones[u] = set(ps)

    email_freq: Dict[str, int] = {}
    phone_freq: Dict[str, int] = {}

    for u, s in per_url_emails.items():
        for e in s:
            email_freq[e.lower()] = email_freq.get(e.lower(), 0) + 1
    for u, s in per_url_phones.items():
        for p in s:
            phone_freq[_norm_phone(p)] = phone_freq.get(_norm_phone(p), 0) + 1

    extracted_email = (data.get('Email') or '').split(',')[0].strip().lower()
    extracted_phone = _norm_phone((data.get('Phone') or '').split(',')[0].strip())

    top_email = max(email_freq.items(), key=lambda x: x[1])[0] if email_freq else ''
    top_email_count = email_freq.get(top_email, 0)

    top_phone = ''
    top_phone_count = 0
    if phone_freq:
        top_phone, top_phone_count = max(phone_freq.items(), key=lambda x: x[1])

    def conf_from_count(cnt: int, total_urls: int) -> float:
        if cnt <= 0:
            return 0.0
        base = 0.5 + min(0.4, (cnt - 1) * 0.15)
        size_bonus = 0.05 if total_urls >= 3 else 0
        return min(0.95, base + size_bonus)

    total_urls = len([h for h in pages.values() if h])

    email_status = 'UNVERIFIED'
    email_conf = 0.0
    chosen_email = extracted_email
    if extracted_email and extracted_email in email_freq and email_freq[extracted_email] >= 2:
        email_conf = conf_from_count(email_freq[extracted_email], total_urls)
        try:
            e_dom = extracted_email.split('@')[-1].lower()
            if site_domain and (e_dom == site_domain or e_dom.endswith('.' + site_domain)):
                email_conf = min(0.98, email_conf + 0.07)
        except Exception:
            pass
        email_status = f"VERIFIED (AI {email_conf:.2f})"
    elif top_email and email_freq[top_email] >= 2:
        chosen_email = top_email
        email_conf = conf_from_count(email_freq[top_email], total_urls)
        try:
            e_dom = top_email.split('@')[-1].lower()
            if site_domain and (e_dom == site_domain or e_dom.endswith('.' + site_domain)):
                email_conf = min(0.98, email_conf + 0.07)
        except Exception:
            pass
        email_status = f"VERIFIED (AI {email_conf:.2f})"
        if not data.get('Email'):
            data['Email'] = top_email

    phone_status = 'UNVERIFIED'
    phone_conf = 0.0
    if extracted_phone and extracted_phone in phone_freq and phone_freq[extracted_phone] >= 2:
        phone_conf = conf_from_count(phone_freq[extracted_phone], total_urls)
        phone_status = f"VERIFIED (AI {phone_conf:.2f})"
    elif top_phone and top_phone_count >= 2:
        phone_conf = conf_from_count(top_phone_count, total_urls)
        phone_status = f"VERIFIED (AI {phone_conf:.2f})"
        if not data.get('Phone'):
            data['Phone'] = top_phone

    status = f"Phone: {phone_status}; Email: {email_status}"
    data['Verification Status'] = status
    return data, status

# ---------- Excel helpers ----------

def to_excel_rows(data_list: List[Dict[str, Any]]) -> pd.DataFrame:
    rows = []
    ts = datetime.utcnow().isoformat()
    for d in data_list:
        row = [d.get(h, "") for h in MOTH_HEADERS[:-1]] + [ts]
        rows.append(row)
    df = pd.DataFrame(rows, columns=MOTH_HEADERS)
    return df


def save_excel_highlight_unverified(df: pd.DataFrame, out_path: str) -> str:
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Results')
        ws = writer.sheets['Results']
        from openpyxl.styles import PatternFill
        red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
        headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        phone_col = headers.get('Phone')
        email_col = headers.get('Email')
        ver_col = headers.get('Verification Status')
        if phone_col and email_col and ver_col:
            for r in range(2, ws.max_row + 1):
                ver = ws.cell(row=r, column=ver_col).value
                if ver and ('FAIL' in str(ver).upper() or 'UNVERIFIED' in str(ver).upper()):
                    ws.cell(row=r, column=phone_col).fill = red_fill
                    ws.cell(row=r, column=email_col).fill = red_fill
    return out_path

# ---------- Search helpers ----------

def search_official_website(company: str, geography: str = "") -> str:
    # If no web search keys, fallback to AI-only domain resolution.
    if not ((GOOGLE_API_KEY and CSE_ID) or BING_SEARCH_KEY):
        return resolve_site_via_ai(company, geography)

    query = f"{company} official site {geography}".strip()
    if GOOGLE_API_KEY and CSE_ID:
        try:
            r = requests.get(
                "https://www.googleapis.com/customsearch/v1",
                params={"q": query, "key": GOOGLE_API_KEY, "cx": CSE_ID}, timeout=12
            )
            js = r.json()
            items = js.get('items') or []
            if items:
                return items[0]['link']
        except Exception as e:
            logger.warning(f"Google CSE search failed: {e}")
    if BING_SEARCH_KEY:
        try:
            r = requests.get(
                "https://api.bing.microsoft.com/v7.0/search",
                headers={"Ocp-Apim-Subscription-Key": BING_SEARCH_KEY},
                params={"q": query, "textDecorations": False, "mkt": "en-US"}, timeout=12
            )
            js = r.json()
            items = ((js.get('webPages') or {}).get('value')) or []
            if items:
                return items[0]['url']
        except Exception as e:
            logger.warning(f"Bing search failed: {e}")
    # Fallback again
    return resolve_site_via_ai(company, geography)

# ---------- Session (multi-run accumulation) ----------
@api_router.post('/session/start')
async def start_session():
    sid = str(uuid.uuid4())
    await db.sessions.insert_one({'session_id': sid, 'items': [], 'created_at': datetime.utcnow().isoformat()})
    return {'session_id': sid}

@api_router.get('/session/{session_id}')
async def get_session(session_id: str):
    s = await db.sessions.find_one({'session_id': session_id})
    if not s:
        raise HTTPException(status_code=404, detail='Session not found')
    return {'session_id': session_id, 'count': len(s.get('items', [])), 'items': s.get('items', [])}

@api_router.get('/session/{session_id}/download')
async def download_session(session_id: str):
    s = await db.sessions.find_one({'session_id': session_id})
    if not s or not s.get('items'):
        raise HTTPException(status_code=404, detail='No items in session')
    out_dir = ROOT_DIR / 'exports'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = str(out_dir / f'{session_id}.xlsx')
    save_excel_highlight_unverified(to_excel_rows(s['items']), out_path)
    return FileResponse(out_path, filename=f'{session_id}.xlsx')

@api_router.post('/session/add/url')
async def session_add_url(session_id: str = Form(...), url: str = Form(...), mode: str = Form(ScrapeMode.REALTIME)):
    s = await db.sessions.find_one({'session_id': session_id})
    if not s:
        raise HTTPException(status_code=404, detail='Session not found')
    pages = crawl_site(url, mode=mode)
    if not pages:
        raise HTTPException(status_code=400, detail='Failed to fetch the site')
    data = gemini_extract(url, pages)
    data, _ = ai_verify_contacts(data, pages, url)
    await db.sessions.update_one({'session_id': session_id}, {'$push': {'items': data}})
    return {'session_id': session_id, 'added': 1, 'count': len((await db.sessions.find_one({'session_id': session_id}))['items'])}

@api_router.post('/session/add/name')
async def session_add_name(session_id: str = Form(...), company_name: str = Form(...), geography: str = Form(''), mode: str = Form(ScrapeMode.REALTIME)):
    s = await db.sessions.find_one({'session_id': session_id})
    if not s:
        raise HTTPException(status_code=404, detail='Session not found')
    site = search_official_website(company_name, geography)
    pages = crawl_site(site, mode=mode)
    if not pages:
        raise HTTPException(status_code=400, detail=f'Failed to fetch site: {site}')
    data = gemini_extract(site, pages)
    data, _ = ai_verify_contacts(data, pages, site)
    await db.sessions.update_one({'session_id': session_id}, {'$push': {'items': data}})
    return {'session_id': session_id, 'added': 1, 'count': len((await db.sessions.find_one({'session_id': session_id}))['items'])}

# ---------- Routes ----------
@api_router.get("/")
async def root():
    return {"message": "Hello World"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_obj = StatusCheck(client_name=input.client_name)
    await db.status_checks.insert_one(status_obj.dict())
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    items = await db.status_checks.find().to_list(1000)
    return [StatusCheck(**it) for it in items]

@api_router.post("/scrape/url", response_model=ScrapeResult)
async def scrape_url(req: ScrapeRequest):
    pages = crawl_site(req.url, mode=req.mode)
    if not pages:
        raise HTTPException(status_code=400, detail='Failed to fetch the site')
    data = gemini_extract(req.url, pages)
    data, _ = ai_verify_contacts(data, pages, req.url)

    job_id = str(uuid.uuid4())
    df = to_excel_rows([data])
    out_dir = ROOT_DIR / 'exports'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = str(out_dir / f'{job_id}.xlsx')
    save_excel_highlight_unverified(df, out_path)

    record = {'job_id': job_id, 'input_url': req.url, 'mode': req.mode, 'data': data, 'created_at': datetime.utcnow().isoformat(), 'excel_path': out_path}
    await db.scrape_jobs.insert_one(record)

    return ScrapeResult(job_id=job_id, status='DONE', data=data, excel_path=out_path, created_at=datetime.utcnow())

@api_router.post("/scrape/name", response_model=ScrapeResult)
async def scrape_by_name(req: NameGeoRequest):
    site = search_official_website(req.company_name, req.geography or "")
    pages = crawl_site(site, mode=req.mode)
    if not pages:
        raise HTTPException(status_code=400, detail=f'Failed to fetch site: {site}')
    data = gemini_extract(site, pages)
    data, _ = ai_verify_contacts(data, pages, site)

    job_id = str(uuid.uuid4())
    df = to_excel_rows([data])
    out_dir = ROOT_DIR / 'exports'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = str(out_dir / f'{job_id}.xlsx')
    save_excel_highlight_unverified(df, out_path)
    record = {'job_id': job_id, 'input_name': req.company_name, 'geo': req.geography, 'mode': req.mode, 'data': data, 'created_at': datetime.utcnow().isoformat(), 'excel_path': out_path}
    await db.scrape_jobs.insert_one(record)
    return ScrapeResult(job_id=job_id, status='DONE', data=data, excel_path=out_path, created_at=datetime.utcnow())

@api_router.post("/bulk/upload")
async def bulk_upload(file: UploadFile = File(...), mode: str = Form(ScrapeMode.REALTIME)):
    content = await file.read()
    name = file.filename or 'upload'
    df: pd.DataFrame
    try:
        if name.lower().endswith('.xlsx') or name.lower().endswith('.xls'):
            df = pd.read_excel(io.BytesIO(content))
        else:
            df = pd.read_csv(io.BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f'Failed to parse file: {e}')

    cols = {c.lower().strip(): c for c in df.columns}
    url_col = cols.get('url') or cols.get('website')
    name_col = cols.get('company') or cols.get('company name') or cols.get('name')
    geo_col = cols.get('geography') or cols.get('country') or cols.get('location')

    results: List[Dict[str, Any]] = []
    errors: List[str] = []

    for i, row in df.iterrows():
        try:
            site = None
            if url_col and isinstance(row[url_col], str) and row[url_col].strip():
                site = str(row[url_col]).strip()
            elif name_col and isinstance(row[name_col], str) and row[name_col].strip():
                site = search_official_website(row[name_col], str(row.get(geo_col) or ''))
            else:
                errors.append(f"Row {i+2}: missing URL or Company name")
                continue

            pages = crawl_site(site, mode=mode)
            if not pages:
                errors.append(f"Row {i+2}: failed to fetch {site}")
                continue
            data = gemini_extract(site, pages)
            data, _ = ai_verify_contacts(data, pages, site)
            results.append(data)
        except Exception as e:
            errors.append(f"Row {i+2}: {e}")

    if not results:
        raise HTTPException(status_code=400, detail=f'No successful rows. Errors: {"; ".join(errors[:5])}')

    job_id = str(uuid.uuid4())
    out_dir = ROOT_DIR / 'exports'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = str(out_dir / f'{job_id}.xlsx')
    save_excel_highlight_unverified(to_excel_rows(results), out_path)

    record = {'job_id': job_id, 'bulk_file': name, 'mode': mode, 'rows': len(results), 'errors': errors, 'created_at': datetime.utcnow().isoformat(), 'excel_path': out_path}
    await db.scrape_jobs.insert_one(record)

    return {"job_id": job_id, "status": "DONE", "rows": len(results), "errors": errors[:10], "download": f"/api/download/{job_id}"}

@api_router.get("/download/{job_id}")
async def download_excel(job_id: str):
    item = await db.scrape_jobs.find_one({'job_id': job_id})
    if not item or not item.get('excel_path') or not Path(item['excel_path']).exists():
        raise HTTPException(status_code=404, detail='File not found')
    return FileResponse(item['excel_path'], filename=f"{job_id}.xlsx")

app.include_router(api_router)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()